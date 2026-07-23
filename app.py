import os
import io
import string
from datetime import datetime
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, session
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__)
app.secret_key = 'kiyya_secret_key_change_this'

# Database Configuration
db_url = os.environ.get('DATABASE_URL', 'sqlite:///students.db')
if db_url.startswith("postgres://"):
    db_url = db_url.replace("postgres://", "postgresql://", 1)

app.config['SQLALCHEMY_DATABASE_URI'] = db_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# Database Models
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)

class Student(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    full_name = db.Column(db.String(100), nullable=False)
    grade = db.Column(db.String(20), nullable=False)
    section = db.Column(db.String(10), default='A')
    phone = db.Column(db.String(20), nullable=True)
    bus_service = db.Column(db.String(50), default='አልፈልግም (No Bus)')
    address = db.Column(db.String(200), nullable=True)
    payment_method = db.Column(db.String(50), nullable=True)
    ft_approval_no = db.Column(db.String(100), nullable=True)
    amount_paid = db.Column(db.Float, default=0.0)
    payment_type = db.Column(db.String(50), nullable=True)
    status = db.Column(db.String(20), default='Pending')
    date_registered = db.Column(db.String(50), default=lambda: datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

class Setting(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    monthly_fee = db.Column(db.Float, default=3000.0)
    term_fee = db.Column(db.Float, default=8000.0)
    bus_fee = db.Column(db.Float, default=1500.0) # <--- NEW SCHOOL BUS FEE SETTING
    class_capacity = db.Column(db.Integer, default=30)
    plan_target_per_grade = db.Column(db.Integer, default=50)
    default_address = db.Column(db.String(200), default="አቃቂ ቃሊቲ ወረዳ 09")

with app.app_context():
    db.create_all()
    if not User.query.filter_by(username='admin').first():
        db.session.add(User(username='admin', password_hash=generate_password_hash('admin123')))
        db.session.commit()
    if not Setting.query.first():
        db.session.add(Setting())
        db.session.commit()

# --- Public Student Registration Page ---
@app.route('/')
def register_page():
    settings = Setting.query.first()
    return render_template('register.html', settings=settings)

@app.route('/add_student', methods=['POST'])
def add_student():
    full_name = request.form.get('full_name')
    grade = request.form.get('grade')
    phone = request.form.get('phone')
    bus_service = request.form.get('bus_service', 'አልፈልግም (No Bus)')
    address = request.form.get('address')
    payment_method = request.form.get('payment_method')
    ft_approval_no = request.form.get('ft_approval_no')
    payment_type = request.form.get('payment_type')
    amount_paid = float(request.form.get('amount_paid', 0))
    
    settings = Setting.query.first()
    capacity = settings.class_capacity if (settings and settings.class_capacity > 0) else 30
    
    existing_count = Student.query.filter_by(grade=grade).count()
    section_index = existing_count // capacity
    assigned_section = string.ascii_uppercase[section_index % 26] 

    new_student = Student(
        full_name=full_name, grade=grade, section=assigned_section,
        phone=phone, bus_service=bus_service, address=address,
        payment_method=payment_method, ft_approval_no=ft_approval_no,
        amount_paid=amount_paid, payment_type=payment_type, status='Pending'
    )
    db.session.add(new_student)
    db.session.commit()
    flash(f'ምዝገባዎ በስኬት ተጠናቋል! የተመደቡበት ክፍል እና ሴክሽን፦ {grade} - {assigned_section}', 'success')
    return redirect(url_for('register_page'))

# --- Auth & Dashboard Routes ---
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        user = User.query.filter_by(username=request.form.get('username')).first()
        if user and check_password_hash(user.password_hash, request.form.get('password')):
            session['user_id'] = user.id
            session['username'] = user.username
            return redirect(url_for('admin_dashboard'))
        flash('የተሳሳተ Username ወይም Password!', 'danger')
    return render_template('login.html')

@app.route('/admin')
def admin_dashboard():
    if 'user_id' not in session: return redirect(url_for('login'))
    
    students = Student.query.all()
    settings = Setting.query.first()
    
    total_expected = 0.0
    total_paid = 0.0
    student_data = []
    
    grades_list = [f"{i}st Grade" if i==1 else f"{i}nd Grade" if i==2 else f"{i}rd Grade" if i==3 else f"{i}th Grade" for i in range(1, 13)]
    grade_stats = {g: 0 for g in grades_list}

    for s in students:
        # Calculate Base Tuition Fee
        base_fee = settings.term_fee if (s.payment_type and 'Term' in s.payment_type) else settings.monthly_fee
        
        # Add Bus Fee if student selected Bus Service
        bus_addon = settings.bus_fee if (s.bus_service and ('Yes' in s.bus_service or 'እፈልጋለሁ' in s.bus_service)) else 0.0
        
        expected = base_fee + bus_addon
        paid = s.amount_paid or 0.0
        remaining = max(0.0, expected - paid)
        
        total_expected += expected
        total_paid += paid
        
        student_data.append({
            'student': s, 'expected': expected, 'paid': paid, 'remaining': remaining, 'bus_addon': bus_addon
        })

        if s.grade in grade_stats:
            grade_stats[s.grade] += 1

    total_unpaid = max(0.0, total_expected - total_paid)
    plan_target = settings.plan_target_per_grade or 50

    return render_template('admin.html', 
                           student_data=student_data, settings=settings,
                           total_expected=total_expected, total_paid=total_paid,
                           total_unpaid=total_unpaid, total_students=len(students),
                           grade_stats=grade_stats, plan_target=plan_target)

@app.route('/approve_student/<int:id>')
def approve_student(id):
    if 'user_id' not in session: return redirect(url_for('login'))
    s = Student.query.get_or_404(id)
    s.status = 'Approved'
    db.session.commit()
    flash(f'የ {s.full_name} ምዝገባ ተረጋግጧል!', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/reject_student/<int:id>')
def reject_student(id):
    if 'user_id' not in session: return redirect(url_for('login'))
    s = Student.query.get_or_404(id)
    s.status = 'Rejected'
    db.session.commit()
    flash(f'የ {s.full_name} ምዝገባ ተሰርዟል!', 'warning')
    return redirect(url_for('admin_dashboard'))

@app.route('/delete_student/<int:id>')
def delete_student(id):
    if 'user_id' not in session: return redirect(url_for('login'))
    s = Student.query.get_or_404(id)
    db.session.delete(s)
    db.session.commit()
    flash('ተማሪው ከዳታቤዝ ተጠርግቧል!', 'danger')
    return redirect(url_for('admin_dashboard'))

@app.route('/update_settings', methods=['POST'])
def update_settings():
    if 'user_id' not in session: return redirect(url_for('login'))
    settings = Setting.query.first()
    settings.monthly_fee = float(request.form.get('monthly_fee', 3000))
    settings.term_fee = float(request.form.get('term_fee', 8000))
    settings.bus_fee = float(request.form.get('bus_fee', 1500)) # <--- UPDATE BUS FEE
    settings.class_capacity = int(request.form.get('class_capacity', 30))
    settings.plan_target_per_grade = int(request.form.get('plan_target_per_grade', 50))
    settings.default_address = request.form.get('default_address')
    db.session.commit()
    flash('ቅንብሮች በስኬት ተቀይረዋል!', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/change_password', methods=['POST'])
def change_password():
    if 'user_id' not in session: return redirect(url_for('login'))
    current_user = User.query.get(session['user_id'])
    if check_password_hash(current_user.password_hash, request.form.get('old_password')):
        current_user.password_hash = generate_password_hash(request.form.get('new_password'))
        db.session.commit()
        flash('የይለፍ ቃልህ በስኬት ተቀይሯል!', 'success')
    else:
        flash('የድሮው የይለፍ ቃል የተሳሳተ ነው!', 'danger')
    return redirect(url_for('admin_dashboard'))

@app.route('/logout')
def logout():
    session.clear()
    flash('ወጥተዋል!', 'info')
    return redirect(url_for('login'))

# --- EXECUTIVE MULTI-SHEET EXPORT FEATURE ---
@app.route('/export_excel')
def export_excel():
    if 'user_id' not in session: return redirect(url_for('login'))
    
    students = Student.query.all()
    settings = Setting.query.first()
    
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    # SHEET 1: Search Dashboard
    ws_dash = wb.create_sheet(title="🔍 Search Dashboard")
    ws_dash.append(["🎓 Wabi School — Executive Dashboard & Analytics"])
    ws_dash.append([])
    ws_dash.append(["🔍 Search by FT / Transaction ID", "", "", "", "📊 Student Registration Plan vs Actual Report"])
    
    target_plan = settings.plan_target_per_grade or 50
    grades_list = [f"{i}st Grade" if i==1 else f"{i}nd Grade" if i==2 else f"{i}rd Grade" if i==3 else f"{i}th Grade" for i in range(1, 13)]
    
    ws_dash.append(["የ FT ቁጥር ያስገቡ፦", "", "", "", "ክፍል (Grade)", "የታቀደ (Plan Target)", "የተመዘገበ (Registered)", "የተመዘገበ % (Completion)"])
    
    for idx, g in enumerate(grades_list):
        reg_count = sum(1 for s in students if s.grade == g)
        pct = (reg_count / target_plan) if target_plan > 0 else 0
        ws_dash.append(["", "", "", "", g, target_plan, reg_count, round(pct, 4)])

    # SHEET 2: All Students
    ws_all = wb.create_sheet(title="All Students")
    headers_all = ["ተ.ቁ (ID)", "ሙሉ ስም", "ስልክ ቁጥር", "ክፍል (Grade)", "ሴክሽን", "ስኩል ባስ", "አድራሻ", "ትራንዛክሽን ቁጥር", "የተከፈለ (ETB)", "የክፍያ ሁኔታ", "የተመዘገበበት ቀን"]
    ws_all.append(headers_all)
    for col_num in range(1, len(headers_all) + 1):
        cell = ws_all.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font

    for s in students:
        ws_all.append([s.id, s.full_name, s.phone, s.grade, s.section, s.bus_service, s.address, s.ft_approval_no, s.amount_paid, s.status, s.date_registered])

    # SHEETS FOR EACH SECTION
    section_groups = {}
    for s in students:
        sec_key = f"{s.grade} - {s.section if s.section else 'A'}"
        if sec_key not in section_groups:
            section_groups[sec_key] = []
        section_groups[sec_key].append(s)

    for sec_name, std_list in sorted(section_groups.items()):
        sheet_title = sec_name.replace(":", "").replace("/", "-")[:30]
        ws = wb.create_sheet(title=sheet_title)
        ws.append(headers_all)
        
        for col_num in range(1, len(headers_all) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font

        for s in std_list:
            ws.append([s.id, s.full_name, s.phone, s.grade, s.section, s.bus_service, s.address, s.ft_approval_no, s.amount_paid, s.status, s.date_registered])

    wb.remove(default_sheet)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name="Wabi_School_Executive_Dashboard.xlsx")

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=True)